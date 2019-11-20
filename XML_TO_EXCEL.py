# -*- coding: utf-8 -*-
"""
Created on Thu Aug 23 13:29:44 2018

@author: Devendra.Tripathi
"""
#import requests
import datetime
import pandas as pd
import os
import json
import openpyxl
from xml.dom import minidom
import requests



Combination=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/Macro File PI.xlsm',sheet_name='Combination')

for Combination_Loop in range(len(Combination)):
    comb=Combination.loc[Combination_Loop,'Combination']
    url='http://ec.europa.eu/eurostat/SDMX/diss-web/rest/data/sts_inpr_m/' +  comb
    print(url)
    try:
        r = requests.get(url)
        with open(os.path.join(os.path.dirname(__file__))+ "/downloaded files/rawdata.xml", 'wb') as f:  
            f.write(r.content)
    except Exception as e:
        print(e)
        
        
    mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '/downloaded files/rawdata.xml')
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    x=2
    sheet.cell(row=1, column=1).value='Unit'
    sheet.cell(row=1, column=2).value='Code'
    sheet.cell(row=1, column=3).value='Geo'
    sheet.cell(row=1, column=4).value='Freq'
    sheet.cell(row=1, column=5).value='Year'
    sheet.cell(row=1, column=6).value='Value'
    
    items = mydoc.getElementsByTagName('generic:Series')
    for elem in items:
        obj1=elem.getElementsByTagName('generic:SeriesKey')[0]        
        strUnit=obj1.getElementsByTagName('generic:Value')[0].attributes['value'].value
        strCode=obj1.getElementsByTagName('generic:Value')[3].attributes['value'].value
        strGeo=obj1.getElementsByTagName('generic:Value')[4].attributes['value'].value
        strFeq=obj1.getElementsByTagName('generic:Value')[5].attributes['value'].value
        
        items1 = elem.getElementsByTagName('generic:Obs')
        for elem1 in items1:
            sheet.cell(row=x, column=1).value=strUnit
            sheet.cell(row=x, column=2).value=strCode
            sheet.cell(row=x, column=3).value=strGeo
            sheet.cell(row=x, column=4).value=strFeq
            sheet.cell(row=x, column=5).value=elem1.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
            sheet.cell(row=x, column=6).value=elem1.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
            x=x+1
    
    wb.save(os.path.join(os.path.dirname(__file__))+ "/downloaded files/rawdata.xlsx")


#        sheet.cell(row=x, column=1).value=elem.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
#        sheet.cell(row=x, column=2).value=elem.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
#        x=x+1
    
#    items = mydoc.getElementsByTagName('generic:Obs')
#    for elem in items:
#        sheet.cell(row=x, column=1).value=elem.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
#        sheet.cell(row=x, column=2).value=elem.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
#        x=x+1
    
#    wb.save(os.path.join(os.path.dirname(__file__))+ "/downloaded file/rawdata.xlsx")


