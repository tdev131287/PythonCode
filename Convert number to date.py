import requests
import openpyxl as xl
import os
import datetime
import pandas as pd



now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
folderName= datestring

folderName=os.path.join(os.path.dirname(__file__)) + '/' + datestring + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)
    

Tickers=pd.read_excel(os.path.join(os.path.dirname(__file__)) + '/'+'Mapping_oilprice.com.xlsx')
for Tickers_Loop in range(len(Tickers)): 
    bent_id=Tickers.loc[Tickers_Loop,'bent id'] 
    filename=Tickers.loc[Tickers_Loop,'name']    
    wb = xl.Workbook()
    sht =wb.active
    sht["A1"].value="Month"
    sht["B1"].value="Value"
    print('Fetching:- ' + filename)
    headers = {
        'sec-fetch-mode': 'cors',
        'origin': 'https://oilprice.com',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'en-IN,en-GB;q=0.9,en-US;q=0.8,en;q=0.7',
        'x-requested-with': 'XMLHttpRequest',
        'cookie': 'oilprice_ci=7dohuqioc7enidk0tt9emuq4fet8rkn3; dmxRegion=false; OX_plg=pm; _wingify_pc_uuid=fa762f51ee184d378a30e4bf00180656; wingify_donot_track_actions=0; __gads=ID=fd4d1da7ac8ce269:T=1573533363:S=ALNI_MZUruM9W5w_LlkrxwvFmXiHZoWWMQ; _omappvp=zcMsJMF9pGeZGiZJgI5X3gs3XFhqzuFP5Eh8It10D0R9vgtMo2H8n5JJhVcV6P8UWXa6S2cL8gg2WZc7k7Yu2dmGutEGLMBk; _ga=GA1.2.121635797.1573533363; _gid=GA1.2.825891892.1573533368; SKpbjs-unifiedid=%7B%22TDID%22%3A%2207487dde-6fd0-49fd-bda2-7b1b34387c96%22%2C%22TDID_LOOKUP%22%3A%22TRUE%22%2C%22TDID_CREATED_AT%22%3A%222019-10-12T04%3A36%3A14%22%7D; SKpbjs-unifiedid_last=Tue%2C%2012%20Nov%202019%2004%3A36%3A14%20GMT; SKpbjs-id5id=%7B%22ID5ID%22%3A%22ID5-ZHMOvAwzdCD-jBhmO-Wuqly7S1yi16LeLFei0ctDoA%22%2C%22ID5ID_CREATED_AT%22%3A%222019-05-11T09%3A09%3A32.629Z%22%2C%22ID5_CONSENT%22%3Atrue%2C%22CASCADE_NEEDED%22%3Atrue%2C%22ID5ID_LOOKUP%22%3Atrue%2C%223PIDS%22%3A%5B%5D%7D; SKpbjs-id5id_last=Tue%2C%2012%20Nov%202019%2004%3A36%3A15%20GMT; __qca=P0-1439342930-1573533367689; _fbp=fb.1.1573533383418.2146925532; __utm_is1=usrsig.73b435cb-44b8-4d35-8668-e6c9d3b4831a.1573533384335; __utm_is2=1573533384335; __utm_is3=544263253678021306267.1573533431215.2; OX_sd=3; _omappvs=1573533434698; mp_f7dc39645bd7dd6a17ccd5e827d4317e_mixpanel=%7B%22distinct_id%22%3A%20%2216e5de610431ff-0184675efb2793-b363e65-ff000-16e5de610442%22%2C%22%24device_id%22%3A%20%2216e5de610431ff-0184675efb2793-b363e65-ff000-16e5de610442%22%2C%22%24initial_referrer%22%3A%20%22https%3A%2F%2Foilprice.com%2F%22%2C%22%24initial_referring_domain%22%3A%20%22oilprice.com%22%7D; _gat_UA-2249023-24=1; AWSALB=A8Qg/LbN1AbcTaZbp2cAxibW61SXnghICf07RCSeqQrmv2E8Vll6m7UhYlxKs8p+uOzVLDNLP2CQZhSXigSb7BPB99Mn8Pk7dtmfEvfEY8Vedd8/FDfnKjwf46ii',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36',
        'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'referer': 'https://oilprice.com/freewidgets/get_oilprices_chart/45/4',
        'authority': 'oilprice.com',
        'sec-fetch-site': 'same-origin',
    }
    
    data = {
      'blend_id': bent_id,
      'period': '7',
      'ci_csrf_token': ''
    }
    
    response = requests.post('https://oilprice.com/freewidgets/json_get_oilprices', headers=headers, data=data)
    loaded_json=response.json()
    
    xrow=2
    for item in loaded_json.keys():
    
        if item=="prices":
            dict1=loaded_json[item]
            for val1 in dict1:
                dt=(((int(val1['time'])/60)/60)/24)+25569
                sht.cell(row=xrow, column=1).value=dt
                sht.cell(row=xrow, column=1).number_format = 'DD-MM-YYYY'
                sht.cell(row=xrow, column=2).value=val1['price']                
                xrow=xrow+1
    sht.column_dimensions['A'].width =18      
    sheet=wb.create_sheet('Details')
    sheet.cell(row=1, column=1).value='Commodity'
    sheet.cell(row=2, column=1).value='Source'
    sheet.cell(row=3,column=1).value='Source Link'
    sheet.cell(row=4,column=1).value='Unit'
    sheet.cell(row=5,column=1).value='Geography'
    
    
    sheet.cell(row=1, column=2).value='Oil Price'
    sheet.cell(row=2, column=2).value='oilprice.com'
    sheet.cell(row=3,column=2).value='https://oilprice.com/oil-price-charts'
    sheet.cell(row=4,column=2).value=''
    sheet.cell(row=5,column=2).value='Global'
    wb.properties.title=filename+'_oilprice.com_Global'
    
    
    wb.save(folderName+ filename+'_oilprice.com_Global.xlsx')            
    wb.close()
    