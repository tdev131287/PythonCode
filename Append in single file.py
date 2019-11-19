import requests
import pandas as pd
from selenium import webdriver
import time
import os
from selenium.webdriver.chrome.options import Options
import glob
import datetime
import openpyxl

now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
#datestring='2019-09-19'
folderName= datestring

directory=os.path.join(os.path.join(os.path.dirname(__file__))+ '/downloaded files/')
os.chdir(directory)
files=glob.glob('*.xls*')
for filename in files:
    os.unlink(filename)
    
    
directory=os.path.join(os.path.join(os.path.dirname(__file__)))
os.chdir(directory)                

#driver = webdriver.Chrome()


chrome_options = webdriver.ChromeOptions()
prefs = {'download.default_directory' : os.path.join(os.path.dirname(__file__))+ '/downloaded files/','profile.default_content_setting_values': {'cookies': 2, 'images': 2, 
                        'plugins': 2, 'popups': 2, 'geolocation': 2, 
                        'notifications': 2, 'auto_select_certificate': 2, 'fullscreen': 2, 
                        'mouselock': 2, 'mixed_script': 2, 'media_stream': 2, 
                        'media_stream_mic': 2, 'media_stream_camera': 2, 'protocol_handlers': 2, 
                        'ppapi_broker': 2, 'automatic_downloads': 2, 'midi_sysex': 2, 
                        'push_messaging': 2, 'ssl_cert_decisions': 2, 'metro_switch_to_desktop': 2, 
                        'protected_media_identifier': 2, 'app_banner': 2, 'site_engagement': 2, 
                        'durable_storage': 2}}
        #    prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)

#all_data = pd.DataFrame() 

all_data=pd.read_excel(os.path.join(os.path.join(os.path.dirname(__file__))+ '/All_Gaziantep Commodity Exchange Market_Turkey.xlsx'))

urlLinks=datestring  
print(urlLinks)

strLink='https://www.gtb.org.tr/gunluk-bulten?tarih=' + str(urlLinks)
print(strLink)   
driver.get(strLink)

strDate=driver.find_element_by_xpath("/html/body/section[1]/div/div/div[1]/div/div[1]").text
strDate=strDate.replace(' - Günlük Bülten Endeksi','')
strDate = datetime.datetime.strptime(strDate, '%d.%m.%Y')
html = driver.page_source
df_list = pd.read_html(html)
try:
    for x in range(1,11):
            df = df_list[-x]
            df['Date']=strDate
            all_data = all_data.append(df)
except:
    pass

try:
    del all_data['Name of the product']
except:
    pass
dfmap=pd.read_excel(os.path.join(os.path.dirname(__file__)) + '/' +'Mapping.xlsx',sheet_name='Sheet1')
result=dfmap.merge(all_data,on='Ürün Adı')  
result.sort_values(by='Date', ascending=False)
result.to_excel(os.path.join(os.path.dirname(__file__))+ '/' + str('All_Gaziantep Commodity Exchange Market_Turkey') +'.xlsx',index=False,index_label=None,columns=None)
driver.quit()


#Creating unique files
directory=os.path.join(os.path.join(os.path.dirname(__file__))+ '/Final files/')
os.chdir(directory)
files=glob.glob('*.xls*')
for filename in files:
    os.unlink(filename)

result=pd.read_excel(os.path.join(os.path.dirname(__file__)) + '/' + 'All_Gaziantep Commodity Exchange Market_Turkey.xlsx',sheet_name='Sheet1')
list1=result['Name of the product'].unique().tolist()
print(list1)

for comname in list1:
    try:
        result1 = result[result['Name of the product'] == comname]
        result1.to_excel(os.path.join(os.path.dirname(__file__)) + '/Final Files/' + comname + '_Gaziantep Commodity Exchange Market_Turkey.xlsx', index=False)
        
        #Creating detail sheet
        wb=openpyxl.load_workbook(os.path.join(os.path.dirname(__file__)) + '/Final Files/' + comname + '_Gaziantep Commodity Exchange Market_Turkey.xlsx')
        sheet=wb.create_sheet('Details')
        
        sheet.cell(row=1, column=1).value='Commodity'
        sheet.cell(row=2, column=1).value='Source'
        sheet.cell(row=3,column=1).value='Source Link'
        sheet.cell(row=4,column=1).value='Geography'
        
        sheet.cell(row=1, column=2).value=comname
        sheet.cell(row=2, column=2).value='Gaziantep Commodity Exchange Market'
        sheet.cell(row=3,column=2).value='https://www.gtb.org.tr'
        sheet.cell(row=4,column=2).value='Turkey'
            
        wb.save(os.path.join(os.path.dirname(__file__)) + '/Final Files/' + comname + '_Gaziantep Commodity Exchange Market_Turkey.xlsx')
        wb.close()
    except:
        pass
    
    



#        #Creating detail sheet
#wb=openpyxl.load_workbook(os.path.join(os.path.dirname(__file__))+ '/' + str('Wheat_Gaziantep Commodity Exchange Market_Turkey') +'.xlsx')
#sheet=wb.create_sheet('Details')
#
#sheet.cell(row=1, column=1).value='Commodity'
#sheet.cell(row=2, column=1).value='Source'
#sheet.cell(row=3,column=1).value='Source Link'
#sheet.cell(row=4,column=1).value='Geography'
#
#sheet.cell(row=1, column=2).value='Wheat'
#sheet.cell(row=2, column=2).value='Gaziantep Commodity Exchange Market'
#sheet.cell(row=3,column=2).value='https://www.gtb.org.tr'
#
#sheet.cell(row=4,column=2).value='Turkey'
#
#wb.save(os.path.join(os.path.dirname(__file__))+ '/' + str('Wheat_Gaziantep Commodity Exchange Market_Turkey') +'.xlsx')
#wb.close()