import os
import openpyxl
import glob
import requests
import json
import pandas as pd
import datetime
import time
import urllib.request
# download file

now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str(now.day)
folderName= datestring

folderName=os.path.join(os.path.dirname(__file__)) + '/' + folderName + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)


URL="https://siocarnes.agroindustria.gob.ar/api/Reportes/GetActividadZona?desde=01/01/2015&hasta=01/08/2019&idZona=-1&idCategoria=1&idSubcategoria=51040401&idRaza=-1"

response = requests.get(URL,verify=False).content

loaded_json = json.loads(response)
#print(loaded_json)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value='fecha'
sheet.cell(row=1, column=2).value='total'
sheet.cell(row=1, column=3).value='promedio'
r=2
for x in loaded_json:
    sheet.cell(row=r, column=1).value=str(x['fecha'])
    sheet.cell(row=r, column=2).value=str(x['total'])
    sheet.cell(row=r, column=3).value=str(x['promedio'])
    r=r+1
  
filename="Livestock - Bovine_" + datestring + ".xlsx"
wb.save(folderName+filename) 


URL="http://monitorsiogranos.magyp.gob.ar/v5_ajax/caracteristicasZonasFechas_min.php?cosas=%7B%0A++%22fechaDesde%22%3A+%2201%2F01%2F2009%22%2C%0A++%22fechaHasta%22%3A+%2207%2F08%2F2019%22%2C%0A++%22IDproducto%22%3A+%2217%22%2C%0A++%22IDzona%22%3A+%2224%22%0A%7D"

response = requests.get(URL,verify=False).content

loaded_json = json.loads(response)
#print(loaded_json)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value='fecha'
sheet.cell(row=1, column=2).value='toneladas'

r=2
for x in loaded_json:
    sheet.cell(row=r, column=1).value=str(x['fecha'])
    sheet.cell(row=r, column=2).value=str(x['toneladas'])
    
    r=r+1
  
filename="Agro - Sunflower_" + datestring + ".xlsx"
wb.save(folderName+filename) 

