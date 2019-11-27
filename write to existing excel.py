import pandas as pd
from openpyxl import load_workbook


df = pd.read_excel('df.xlsx',sheetname='Sheet1') 

#
def append_df_to_excel(filename, df, sheet_name='Sheet', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
   


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, index=False,startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
if __name__ == '__main__':
    df = pd.read_excel('df.xlsx',sheetname='Sheet1') 
    append_df_to_excel('Existing_File.xlsx', df, header=False,sheet_name='Sheet')