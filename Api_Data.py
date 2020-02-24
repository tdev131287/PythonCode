# -*- coding: utf-8 -*-
"""
Created on Thu Aug 23 13:29:44 2018

@author: Devendra.Tripathi
"""
import requests
import datetime
import pandas as pd
import os
import json
import openpyxl
import openpyxl as xl
import glob
from xml.dom import minidom


class ApiData:

    def downloadfile(self,url,fname):
        try:
            r = requests.get(url)
            with open(fname, 'wb') as f:  
                f.write(r.content)
        except Exception as e:
            print(e)
    
    
    def BLSExportData(self,list1):
        now = datetime.datetime.now()
#        list1=['PCU21111221111211','CUSR0000SAA2']
        strEndYEar =str(now.year) 
        #import prettytable
        headers = {'Content-type': 'application/json','registrationkey' : '92d212c02ffa48efb63a03cf66083200'}
        data = json.dumps({"seriesid": list1,"startyear":"2010", "endyear":"2019","registrationkey":"92d212c02ffa48efb63a03cf66083200"})
        p = requests.post('https://api.bls.gov/publicAPI/v2/timeseries/data/', data=data, headers=headers)
        json_data = json.loads(p.text)
        wb = openpyxl.Workbook()
        sheet = wb.active
        x=2
        sheet.cell(row=1, column=1).value='ID'
        sheet.cell(row=1, column=2).value='year'
        sheet.cell(row=1, column=3).value='Period'
        sheet.cell(row=1, column=4).value='PeriodName'
        sheet.cell(row=1, column=5).value='Value'
        
        
        for series in json_data['Results']['series']:
        #    x=prettytable.PrettyTable(["series id","year","period","value","footnotes"])
            seriesId = series['seriesID']    
            for item in series['data']:
                year = item['year']
                period = item['period']
                periodName=item['periodName']
                value = item['value']
                sheet.cell(row=x, column=1).value=seriesId
                sheet.cell(row=x, column=2).value=year
                sheet.cell(row=x, column=3).value=period
                sheet.cell(row=x, column=4).value=periodName
                sheet.cell(row=x, column=5).value=value        
#                footnotes=""
                x=x+1
        wb.save(os.path.join(os.path.dirname(__file__))+ "/Downloaded/blsAllData.xlsx")
        #Read Data from XML ---
    def EIA_Data(self,url,filename):
        
        p = requests.get(url)
        x= p.json()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1).value='Date'
        sheet.cell(row=1, column=2).value='Value'
        r=2
        for i in x['series'][0]['data']:
#            print(i[0])
            year = i[0]
            xvalue = i[1]
#            print( str(year) +" - -"+ str(xvalue))
            
            sheet.cell(row=r, column=1).value=year
            sheet.cell(row=r, column=2).value=xvalue
            r=r+1
        wb.save(os.path.join(os.path.dirname(__file__))+ "/Downloaded/"+filename)
        
    def Download_USDA_Data(self,comCode,strYear):
        url="https://apps.fas.usda.gov/PSDOnlineDataServices/api/CommodityData/GetCommodityDataByYear?commodityCode="+comCode+"&marketYear=" + strYear
        headers={'API_KEY' : '7E8BF00A-A7AA-4705-B23F-970A39C86C95'}
        resp = requests.get(url,headers=headers)
        wb = openpyxl.Workbook()
        sheet = wb.active
        x=2
        sheet.cell(row=1, column=1).value='CommodityCode'
        sheet.cell(row=1, column=2).value='CommodityDescription'
        sheet.cell(row=1, column=3).value='CountryCode'
        sheet.cell(row=1, column=4).value='CountryName'
        sheet.cell(row=1, column=5).value='MarketYear'
        sheet.cell(row=1, column=6).value='CalendarYear'
        sheet.cell(row=1, column=7).value='Month'
        sheet.cell(row=1, column=8).value='AttributeId'
        sheet.cell(row=1, column=9).value='AttributeDescription'
        sheet.cell(row=1, column=10).value='UnitId'
        sheet.cell(row=1, column=11).value='UnitDescription'  
        sheet.cell(row=1, column=12).value='Value'
        
        try:
            for i in resp.json():
                try:
                    sheet.cell(row=x, column=1).value=i['CommodityCode']
                except:
                    sheet.cell(row=x, column=1).value=None        
                try:
                    sheet.cell(row=x, column=2).value=i['CommodityDescription']
                except:
                    sheet.cell(row=x, column=2).value=None
                try:
                    sheet.cell(row=x, column=3).value=i['CountryCode']
                except:
                    sheet.cell(row=x, column=3).value=None
                try:
                    sheet.cell(row=x, column=4).value=i['CountryName']
                except:
                    sheet.cell(row=x, column=4).value=None
                try:
                    sheet.cell(row=x, column=5).value=i['MarketYear']
                except:
                    sheet.cell(row=x, column=5).value=None
                try:            
                    sheet.cell(row=x, column=6).value=i['CalendarYear']
                except:
                    sheet.cell(row=x, column=6).value=None
                try:
                    sheet.cell(row=x, column=7).value=i['Month']
                except:
                    sheet.cell(row=x, column=7).value=None
                try:  
                     sheet.cell(row=x, column=8).value=i['AttributeId']
                except:
                    sheet.cell(row=x, column=8).value=None
                try:
                    sheet.cell(row=x, column=9).value=i['AttributeDescription']
                except:
                    sheet.cell(row=x, column=9).value=None
                try:
                    sheet.cell(row=x, column=10).value=i['UnitId']
                except:
                    sheet.cell(row=x, column=10).value=None
                try:
                    sheet.cell(row=x, column=11).value=i['UnitDescription']    
                except:
                    sheet.cell(row=x, column=11).value=None
                try:
                    sheet.cell(row=x, column=12).value=i['Value']
                except:
                    sheet.cell(row=x, column=12).value=None
                    
                x=x+1
        except:
            print('data not found '+ comCode+"_"+strYear)
        wb.save(os.path.join(os.path.dirname(__file__))+ '/USDA/downloaded files/' + comCode+"_"+strYear+".xlsx")
                

    def Eurostat_CPI(self):
        
        fpath ='file:///' + os.path.join(os.path.dirname(__file__)) + '/Eurostat/CPI/Macro File PI.xlsm'
        
        print(fpath)
        Combination=pd.read_excel(fpath,sheet_name='Combination')
        print(Combination)
        for Combination_Loop in range(len(Combination)):
            comb=Combination.loc[Combination_Loop,'Combination']
            url='http://ec.europa.eu/eurostat/SDMX/diss-web/rest/data/ei_cphi_m/' +  comb
            print(url)
            print('Call this function')
            try:
                r = requests.get(url)
                with open(os.path.join(os.path.dirname(__file__))+ "\\Eurostat\\CPI\\downloaded files\\rawdata.xml", 'wb') as f:  
                    f.write(r.content)
            except Exception as e:
                print(e)
            mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '\\Eurostat\\CPI/downloaded files\\rawdata.xml')
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            x=2
            sheet.cell(row=1, column=1).value='Unit'
            sheet.cell(row=1, column=2).value='Code'
            sheet.cell(row=1, column=3).value='Geo'
            sheet.cell(row=1, column=4).value='Freq'
            sheet.cell(row=1, column=5).value='Year'
            sheet.cell(row=1, column=6).value='Value'
            
            items = mydoc.getElementsByTagName('generic:Series')
            for elem in items:
                obj1=elem.getElementsByTagName('generic:SeriesKey')[0]        
                strUnit=obj1.getElementsByTagName('generic:Value')[0].attributes['value'].value
                strCode=obj1.getElementsByTagName('generic:Value')[1].attributes['value'].value
                strGeo=obj1.getElementsByTagName('generic:Value')[3].attributes['value'].value
                strFeq=obj1.getElementsByTagName('generic:Value')[4].attributes['value'].value
                
                items1 = elem.getElementsByTagName('generic:Obs')
                for elem1 in items1:
                    sheet.cell(row=x, column=1).value=strUnit
                    sheet.cell(row=x, column=2).value=strCode
                    sheet.cell(row=x, column=3).value=strGeo
                    sheet.cell(row=x, column=4).value=strFeq
                    sheet.cell(row=x, column=5).value=elem1.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
                    sheet.cell(row=x, column=6).value=elem1.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
                    x=x+1
            
            wb.save(os.path.join(os.path.dirname(__file__))+ "\\Eurostat\\CPI\\downloaded files\\rawdata.xlsx")
        
    def Eurostat_PPI(self):
        Combination=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/Eurostat/PPI/Macro File PI.xlsm',sheet_name='Combination')
        for Combination_Loop in range(len(Combination)):
            comb=Combination.loc[Combination_Loop,'Combination']
            url='http://ec.europa.eu/eurostat/SDMX/diss-web/rest/data/sts_inpp_m/' +  comb
            print(url)
            try:
                r = requests.get(url)
                with open(os.path.join(os.path.dirname(__file__))+ "/Eurostat/PPI/downloaded files/rawdata.xml", 'wb') as f:  
                    f.write(r.content)
            except Exception as e:
                print(e)
            mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '/Eurostat/PPI/downloaded files/rawdata.xml')
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            x=2
            sheet.cell(row=1, column=1).value='Unit'
            sheet.cell(row=1, column=2).value='Code'
            sheet.cell(row=1, column=3).value='Geo'
            sheet.cell(row=1, column=4).value='Freq'
            sheet.cell(row=1, column=5).value='Year'
            sheet.cell(row=1, column=6).value='Value'
            
            items = mydoc.getElementsByTagName('generic:Series')
            for elem in items:
                obj1=elem.getElementsByTagName('generic:SeriesKey')[0]        
                strUnit=obj1.getElementsByTagName('generic:Value')[0].attributes['value'].value
                strCode=obj1.getElementsByTagName('generic:Value')[2].attributes['value'].value
                strGeo=obj1.getElementsByTagName('generic:Value')[4].attributes['value'].value
                strFeq=obj1.getElementsByTagName('generic:Value')[5].attributes['value'].value
                
                items1 = elem.getElementsByTagName('generic:Obs')
                for elem1 in items1:
                    sheet.cell(row=x, column=1).value=strUnit
                    sheet.cell(row=x, column=2).value=strCode
                    sheet.cell(row=x, column=3).value=strGeo
                    sheet.cell(row=x, column=4).value=strFeq
                    sheet.cell(row=x, column=5).value=elem1.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
                    sheet.cell(row=x, column=6).value=elem1.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
                    x=x+1
            wb.save(os.path.join(os.path.dirname(__file__))+ "/Eurostat/PPI/downloaded files/rawdata.xlsx")
        
    def ProductionIndustry(self):
        Combination=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/Eurostat/Production in industry/Macro File PI.xlsm',sheet_name='Combination')
        
        for Combination_Loop in range(len(Combination)):
            comb=Combination.loc[Combination_Loop,'Combination']
            url='http://ec.europa.eu/eurostat/SDMX/diss-web/rest/data/sts_inpr_m/' +  comb
            print(url)
            try:
                r = requests.get(url)
                with open(os.path.join(os.path.dirname(__file__))+ "/Eurostat/Production in industry/downloaded files/rawdata.xml", 'wb') as f:  
                    f.write(r.content)
            except Exception as e:
                print(e)
            mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '/Eurostat/Production in industry/downloaded files/rawdata.xml')
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            x=2
            sheet.cell(row=1, column=1).value='Unit'
            sheet.cell(row=1, column=2).value='Code'
            sheet.cell(row=1, column=3).value='Geo'
            sheet.cell(row=1, column=4).value='Freq'
            sheet.cell(row=1, column=5).value='Year'
            sheet.cell(row=1, column=6).value='Value'
            
            items = mydoc.getElementsByTagName('generic:Series')
            for elem in items:
                obj1=elem.getElementsByTagName('generic:SeriesKey')[0]        
                strUnit=obj1.getElementsByTagName('generic:Value')[0].attributes['value'].value
                strCode=obj1.getElementsByTagName('generic:Value')[3].attributes['value'].value
                strGeo=obj1.getElementsByTagName('generic:Value')[4].attributes['value'].value
                strFeq=obj1.getElementsByTagName('generic:Value')[5].attributes['value'].value
                
                items1 = elem.getElementsByTagName('generic:Obs')
                for elem1 in items1:
                    sheet.cell(row=x, column=1).value=strUnit
                    sheet.cell(row=x, column=2).value=strCode
                    sheet.cell(row=x, column=3).value=strGeo
                    sheet.cell(row=x, column=4).value=strFeq
                    sheet.cell(row=x, column=5).value=elem1.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
                    sheet.cell(row=x, column=6).value=elem1.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
                    x=x+1
            
            wb.save(os.path.join(os.path.dirname(__file__))+ "/Eurostat/Production in industry/downloaded files/rawdata.xlsx")

    def ProductionIndustryConstruction(self):
        Combination=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/Eurostat/Production in industry - Construction/Macro File PI_Construction.xlsm',sheet_name='Combination')
        
        for Combination_Loop in range(len(Combination)):
            comb=Combination.loc[Combination_Loop,'Combination']
            url='http://ec.europa.eu/eurostat/SDMX/diss-web/rest/data/sts_copr_m/' +  comb
            print(url)
            try:
                r = requests.get(url)
                with open(os.path.join(os.path.dirname(__file__))+ "/Eurostat/Production in industry - Construction/downloaded files/rawdata.xml", 'wb') as f:  
                    f.write(r.content)
            except Exception as e:
                print(e)
            mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '/Eurostat/Production in industry - Construction/downloaded files/rawdata.xml')
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            x=2
            sheet.cell(row=1, column=1).value='S_ADJ'
            sheet.cell(row=1, column=2).value='Code'
            sheet.cell(row=1, column=3).value='Geo'
            sheet.cell(row=1, column=4).value='Freq'
            sheet.cell(row=1, column=5).value='Year'
            sheet.cell(row=1, column=6).value='Value'
            
            items = mydoc.getElementsByTagName('generic:Series')
            for elem in items:
                obj1=elem.getElementsByTagName('generic:SeriesKey')[0]        
                strUnit=obj1.getElementsByTagName('generic:Value')[2].attributes['value'].value
                strCode=obj1.getElementsByTagName('generic:Value')[3].attributes['value'].value
                strGeo=obj1.getElementsByTagName('generic:Value')[4].attributes['value'].value
                strFeq=obj1.getElementsByTagName('generic:Value')[5].attributes['value'].value
                
                items1 = elem.getElementsByTagName('generic:Obs')
                for elem1 in items1:
                    sheet.cell(row=x, column=1).value=strUnit
                    sheet.cell(row=x, column=2).value=strCode
                    sheet.cell(row=x, column=3).value=strGeo
                    sheet.cell(row=x, column=4).value=strFeq
                    sheet.cell(row=x, column=5).value=elem1.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
                    sheet.cell(row=x, column=6).value=elem1.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
                    x=x+1
            
            wb.save(os.path.join(os.path.dirname(__file__))+ "/Eurostat/Production in industry - Construction/downloaded files/rawdata.xlsx")
            
    def ProductionIndustryDomesticMarket(self):
        Combination=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/Eurostat/Production in industry Domestic market/Macro File PI Domestic Market.xlsm',sheet_name='Combination')
        
        for Combination_Loop in range(len(Combination)):
            comb=Combination.loc[Combination_Loop,'Combination']
            url='http://ec.europa.eu/eurostat/SDMX/diss-web/rest/data/sts_inppd_m/' +  comb
            print(url)
            try:
                r = requests.get(url)
                with open(os.path.join(os.path.dirname(__file__))+ "/Eurostat/Production in industry Domestic market/downloaded files/rawdata.xml", 'wb') as f:  
                    f.write(r.content)
            except Exception as e:
                print(e)
            mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '/Eurostat/Production in industry Domestic market/downloaded files/rawdata.xml')
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            x=2
            sheet.cell(row=1, column=1).value='Unit'
            sheet.cell(row=1, column=2).value='Code'
            sheet.cell(row=1, column=3).value='Geo'
            sheet.cell(row=1, column=4).value='Freq'
            sheet.cell(row=1, column=5).value='Year'
            sheet.cell(row=1, column=6).value='Value'
            
            items = mydoc.getElementsByTagName('generic:Series')
            for elem in items:
                obj1=elem.getElementsByTagName('generic:SeriesKey')[0]        
                strUnit=obj1.getElementsByTagName('generic:Value')[0].attributes['value'].value
                strCode=obj1.getElementsByTagName('generic:Value')[2].attributes['value'].value
                strGeo=obj1.getElementsByTagName('generic:Value')[4].attributes['value'].value
                strFeq=obj1.getElementsByTagName('generic:Value')[5].attributes['value'].value
                
                items1 = elem.getElementsByTagName('generic:Obs')
                for elem1 in items1:
                    sheet.cell(row=x, column=1).value=strUnit
                    sheet.cell(row=x, column=2).value=strCode
                    sheet.cell(row=x, column=3).value=strGeo
                    sheet.cell(row=x, column=4).value=strFeq
                    sheet.cell(row=x, column=5).value=elem1.getElementsByTagName('generic:ObsDimension')[0].attributes['value'].value
                    sheet.cell(row=x, column=6).value=elem1.getElementsByTagName('generic:ObsValue')[0].attributes['value'].value
                    x=x+1
            
            wb.save(os.path.join(os.path.dirname(__file__))+ "/Eurostat/Production in industry Domestic market/downloaded files/rawdata.xlsx")
            
    def BEA_PersonalConsumation(self):
        url='https://apps.bea.gov/api/data/?&UserID=182DBCCE-8012-4F7D-B9A9-CEDAB4597A8D%20&method=GetData&DataSetName=NIUnderlyingDetail&TableName=U20305&Frequency=M&Year=ALL&ResultFormat=xml'
        print(url)
        try:
            r = requests.get(url)
            with open(os.path.join(os.path.dirname(__file__))+ "/BEA/Personal Consumption Expenditures/downloaded files/rawdata.xml", 'wb') as f:  
                f.write(r.content)
        except Exception as e:
            print(e)
        mydoc = minidom.parse(os.path.join(os.path.dirname(__file__)) +  '/BEA/Personal Consumption Expenditures/downloaded files/rawdata.xml')
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        x=2
        sheet.cell(row=1, column=1).value='Time'
        sheet.cell(row=1, column=2).value='Value'
        items = mydoc.getElementsByTagName('Results')
        for elem in items:
            
            items1 = elem.getElementsByTagName('Data')
            for elem1 in items1:
        
                if elem1.attributes['SeriesCode'].value=='DPCERC':
                    timePerirod=elem1.attributes['TimePeriod'].value
                    year=timePerirod[:4]
                    Month=timePerirod[-2:]
                    sheet.cell(row=x, column=1).value=datetime.date(int(year), int(Month), 1)
                    xitem =elem1.attributes['DataValue'].value.replace(",","")
#                    print('Before Value' + elem1.attributes['DataValue'].value +' After Value ' + xitem)
                    sheet.cell(row=x, column=2).value=int(xitem)/1000
                    x=x+1
        
        wb.save(os.path.join(os.path.dirname(__file__))+ "/BEA/Personal Consumption Expenditures/downloaded files/rawdata.xlsx")
        
    def USCensusData(self):
        
        directory=os.path.join(os.path.dirname(__file__))+ '/US Census/downloaded files/'
        os.chdir(directory)
        files=glob.glob('*.xlsx')
        for filename in files:
            os.unlink(filename)
            
        os.chdir(os.path.join(os.path.dirname(__file__)))
        dwb = xl.load_workbook(os.path.join(os.path.dirname(__file__))+'/US Census/US Census - HS Codes.xlsm', read_only=False, keep_vba=True)
        dsht=dwb.get_sheet_by_name('Codes')
#        mainCode = dsht.cell(row=2, column=2).value
#        mainComm = dsht.cell(row=2, column=3).value                    
        
        Tickers=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/US Census/US Census - HS Codes.xlsm',sheet_name='Codes')
        
#        row1 = 1
#        row2 = 1
        
        try:
            workbook1=xl.load_workbook(os.path.join(os.path.dirname(__file__))+ "/US Census/downloaded files/"+ "Raw_Data.xlsx")
        except:
            workbook1=xl.Workbook()
                
        worksheet=workbook1.active
        
        for Tickers_Loop in range(len(Tickers)):

            filename=Tickers.loc[Tickers_Loop,'SelCode']
            

            
            
            code=Tickers.loc[Tickers_Loop,'Codes']
            code=code.replace(".","")
            Descrition=Tickers.loc[Tickers_Loop,'Descrition']
            
            iurl='https://api.census.gov/data/timeseries/intltrade/imports/hs?get=I_COMMODITY_SDESC,UNIT_QY1,UNIT_QY2,CON_QY1_MO,CON_QY2_MO,CTY_CODE,CTY_NAME&time=from2013&I_COMMODITY=' + code + '&key=55dd40fea499f3632468ad3542fa3a94ba5bd6de'
            
            print(iurl)
            resp = requests.get(iurl)
            
            if resp.text== '':
                print('No Data Found')
            else:

                row1=worksheet.max_row

    
                for i in resp.json():
                    worksheet.cell(row=row1, column=8).value=i[0]
                    worksheet.cell(row=row1, column=2).value=i[1]
                    worksheet.cell(row=row1, column=3).value=i[2]
                    worksheet.cell(row=row1, column=4).value=i[3]
                    worksheet.cell(row=row1, column=5).value=i[4]
                    worksheet.cell(row=row1, column=6).value=i[5]
                    worksheet.cell(row=row1, column=7).value=i[6]
                    
                    try:
                        timePerirod=i[7]
                        year=timePerirod[:4]
                        Month=timePerirod[-2:]
                        worksheet.cell(row=row1, column=1).value=datetime.date(int(year), int(Month), 1)
                    except:
                        worksheet.cell(row=row1, column=1).value=i[7]
                        
                    worksheet.cell(row=row1, column=9).value=i[8]
                    worksheet.cell(row=row1, column=10).value="Import"
                    worksheet.cell(row=row1, column=11).value=filename
                    worksheet.cell(row=row1, column=12).value=Descrition
                    
                    row1 += 1
            
            

            eurl='https://api.census.gov/data/timeseries/intltrade/exports/hs?get=E_COMMODITY_SDESC,UNIT_QY1,UNIT_QY2,QTY_1_MO,QTY_2_MO,CTY_CODE,CTY_NAME&time=from2013&E_COMMODITY=' + code + '&key=55dd40fea499f3632468ad3542fa3a94ba5bd6de'
            
            print(eurl)
            resp = requests.get(eurl)
            
            if resp.text== '':
                print('No Data Found')
            else:

                row1=worksheet.max_row
#                row1 = 1
#                col = 1
    
                for i in resp.json():
                    worksheet.cell(row=row1, column=8).value=i[0]
                    worksheet.cell(row=row1, column=2).value=i[1]
                    worksheet.cell(row=row1, column=3).value=i[2]
                    worksheet.cell(row=row1, column=4).value=i[3]
                    worksheet.cell(row=row1, column=5).value=i[4]
                    worksheet.cell(row=row1, column=6).value=i[5]
                    worksheet.cell(row=row1, column=7).value=i[6]
                    
                    try:
                        timePerirod=i[7]
                        year=timePerirod[:4]
                        Month=timePerirod[-2:]
                        worksheet.cell(row=row1, column=1).value=datetime.date(int(year), int(Month), 1)
                    except:
                        worksheet.cell(row=row1, column=1).value=i[7]
                        
                    worksheet.cell(row=row1, column=9).value=i[8]
                    worksheet.cell(row=row1, column=10).value="Export"
                    worksheet.cell(row=row1, column=11).value=filename
                    worksheet.cell(row=row1, column=12).value=Descrition
                    row1 += 1
                    
#        workbook1.save(os.path.join(os.path.dirname(__file__))+ "/downloaded files/"+ str(filename)+"_Export.xlsx")
        workbook1.save(os.path.join(os.path.dirname(__file__))+ "/US Census/downloaded files/"+ "Raw_Data.xlsx")
        print("Data downlaoded successfully!!")
        

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ C R E A T E    -   O B JE C T @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
        
apiObj = ApiData()
now = datetime.datetime.now()
stYear=str(now.year) 

#apiObj.USCensusData()
#
##print(now.year)
#urlx1 = 'http://api.eia.gov/series/?api_key=ebdae64909360fd6c4caca4850086a5c&series_id=STEO.BREPUUS.M'
#urlx2 = 'http://api.eia.gov/series/?api_key=ebdae64909360fd6c4caca4850086a5c&series_id=STEO.EOPRPUS.M'
#
#apiObj.EIA_Data(urlx1,"EIA_52.xlsx")
#apiObj.EIA_Data(urlx2,"EIA_69.xlsx")
#
#commoditylist=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/CommodityCode.xlsx',sheet_name='Codelist')
#codeList=[]
#for comCode in range(len(commoditylist)):
#    code=commoditylist.loc[comCode,'series_id']
#    codeList.append(code)
# 
#apiObj.BLSExportData(codeList)
## - Download PI Data -
#ipurl='https://www.federalreserve.gov/datadownload/Output.aspx?rel=G17&series=c8dbcf9065357edb5b7b1b66b4eae6e4&lastobs=&from=01/01/1975&to=12/31/'+str(stYear)+'&filetype=spreadsheetml&label=include&layout=seriescolumn'
#path=os.path.join(os.path.dirname(__file__))+ "/Downloaded/IP_Pricess.xls"
#apiObj.downloadfile(ipurl,path)

## - -------------------------------------- Done
#apiObj.Eurostat_CPI()
#
## - --------------------------------------
#apiObj.Eurostat_PPI()
#
## - --------------------------------------
#apiObj.ProductionIndustry()

## - --------------------------------------
#apiObj.ProductionIndustryConstruction()
#
## - --------------------------------------
#apiObj.ProductionIndustryDomesticMarket()

## - --------------------------------------
#apiObj.BEA_PersonalConsumation()

# - --------------------------------------
years=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/USDA/commodities_code.xlsx',sheet_name='Year',type={"Year": str})
Tickers1=pd.read_excel('file:///' + os.path.join(os.path.dirname(__file__)) + '/USDA/commodities_code.xlsx',sheet_name='Commodity Code',type={"Commodity Code": str})
for Tickers_Loop in range(len(Tickers1)):     
    comCode=Tickers1.loc[Tickers_Loop,'Commodity Code']
#    print(comCode)
    for years_Loop in range(len(years)):
        year=years.loc[years_Loop,'Year']
#        print(year)
        apiObj.Download_USDA_Data(comCode,year)
        
# - --------------------------------------

