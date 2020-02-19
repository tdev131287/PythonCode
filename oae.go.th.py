# -*- coding: utf-8 -*-
"""
Created on Fri Oct 11 15:11:16 2019

@author: Devendra.Tripathi
"""

from selenium import webdriver
import datetime
from TSCLibrary import FolderSystem
import os
import time
import pandas as pd
from pynput.keyboard import Key, Controller
import openpyxl as xl
import win32com.client as win32
import glob

keyboard = Controller()

now= datetime.datetime.now()
xdate =str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
fpath=os.path.join(os.path.dirname(__file__))
filepath =FolderSystem.CreateFolder(fpath)
filepath=filepath.replace("/","\\")

tmppath=os.path.join(os.path.dirname(__file__))+'/TEMP/'
if not os.path.exists(tmppath):
    os.mkdir(tmppath)
tmppath=tmppath.replace("/","\\")  
chrome_options = webdriver.ChromeOptions()
#prefs = {'download.default_directory' : tmppath}
prefs = {
  "translate_whitelists": {"th":"en"},
  "translate":{"enabled":"true"}
}
chrome_options.add_experimental_option('prefs', prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)

df = pd.read_excel(os.path.join(os.path.dirname(__file__)) +'/Mapping.xlsx')
for xline in range(len(df)):
    try:
        link = df.loc[xline,'Links']
        commodity = df.loc[xline,'ProductName']
        
    
        driver.get(link)
        keyboard.press(Key.end)
        time.sleep(10)
        html = driver.page_source
#        print('Url_____ : '+link +'Commodity : '+commodity)
        df_list = pd.read_html(html)
        pgdf = df_list[-1]
#        print('Number of page Count  ' +str(pgdf))
        pgCount = pgdf.iloc[0,1].replace("Page 1 of ","")
        
        
        mydf = pd.DataFrame(columns=['Product Name', 'Week', 'Date','Price','(+/-)','Unit','Reference'])
        pg=1
        try:
            print('Actual of page Count  ' +str(pgCount))
            while pg<=int(pgCount):
                try:
                    
                    nxturl = link + "&page="+str(pg)
                    driver.get(nxturl)
                    
                    time.sleep(5)
                    keyboard.press(Key.end)
                    time.sleep(5)
                    html = driver.page_source
                    
                    df_list = pd.read_html(html,header=0)
                    df1 = df_list[-2]
#                    print(df1.head[20])
                    print('Before Running - ' +  link + '---- '+str(pg))
                    df1['Product Name']=df1.at[0,'Product Name']
                    mydf=mydf.append(df1)
                    print('After Running - ' +  link + '---- '+str(pg))
                    pg=pg+1
                except:
                    print('Find Error in pg number :- ' + str(pg))
        #
            
            mydf.to_excel(filepath+commodity +'_CPF Feed Marekting bureau_Thailand.xlsx',index=False)
            
            wb = xl.load_workbook(filename=filepath+commodity +'_CPF Feed Marekting bureau_Thailand.xlsx')
            sheet=wb.create_sheet('Details')
            sheet.cell(row=1, column=1).value='Commodity'
            sheet.cell(row=2, column=1).value='Source'
            sheet.cell(row=3,column=1).value='Source Link'
            sheet.cell(row=4,column=1).value='Unit'
            sheet.cell(row=5,column=1).value='Geography'
            
            
            sheet.cell(row=1, column=2).value=commodity
            sheet.cell(row=2, column=2).value='CPF Feed Marekting bureau'
            sheet.cell(row=3,column=2).value= link
            sheet.cell(row=4,column=2).value='NA'
            sheet.cell(row=5,column=2).value='Thailand'
            wb.properties.title=commodity +'_CPF Feed Marekting bureau_Thailand'
            wb.save(filepath+commodity +'_CPF Feed Marekting bureau_Thailand.xlsx')
            wb.close()
#            break
        except:
            print('Find error')
    except:
        print('Get error in read url')


    
driver.quit()

#outlook = win32.Dispatch('outlook.application')
#mail = outlook.CreateItem(0)
##mail.To = 'hemant.saigal@thesmartcube.com;maksood.alam@thesmartcube.com;automation@thesmartcube.com'
#mail.To = 'devendra.tripathi@thesmartcube.com;maksood.alam@thesmartcube.com;hemant.saigal@thesmartcube.com'
#mail.Subject = 'Commodity Prices:- charting.kentgroupltd.com'
#mail.Body = 'Pleae find the attached files'
#
#directory=filepath
#os.chdir(directory)
#files=glob.glob('*.xlsx')
#for filename in files:
#    print(os.path.join(os.path.dirname(__file__))+ filename)
##    attachment='E:\\Hemant\\a3m-asso.fr\\' + filename
#    attachment=filepath + filename
#    mail.Attachments.Add(attachment)
#mail.Send()
