import pandas as pd
import os
import datetime
import openpyxl
import requests
import re

#now = datetime.datetime.now()
now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
folderName= datestring

folderName=os.path.join(os.path.dirname(__file__)) + '/' + 'Downloaded files' + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)
    
url = "https://ahdb.org.uk/UKFeedPricesExport/API?id=23007&currency=&timescale=MWeek;y&startDate=2010-01-01&endDate="+datestring
resp = requests.get(url)
with open(os.path.join(os.path.dirname(__file__)) + '/'+ 'Data All.xlsx', 'wb') as output:
    output.write(resp.content)
df=pd.read_excel(os.path.join(os.path.dirname(__file__)) + '/'+ 'Data All.xlsx',header=2)
#df=pd.read_excel(folderName+ 'test.xlsx')
#df=df.drop(df.index[len(df)-3])
df=df[:-3]
columns = list(df)
for x in range(2,len(columns)):
    print('Creating ' + str(columns[x]))
    cols = [0,1,x]
    df1 = df[df.columns[cols]]
#    print(columns[x])
#    print(str(x))
#    header = ["Date", "Delivery Month", columns[x]]
    name=str(columns[x]).replace('%' ,' per ')
    name=re.sub('[^A-Za-z0-9 ]+', '', name)
    
    df1.to_excel(folderName+ name+'_AHDB - Ferilizers and Feed_United Kingdom.xlsx',index=False)
    wb=openpyxl.load_workbook(filename=folderName +  name+'_AHDB - Ferilizers and Feed_United Kingdom.xlsx')
    sheet=wb.create_sheet('Details')
    sheet.cell(row=1, column=1).value='Commodity'
    sheet.cell(row=2, column=1).value='Source'
    sheet.cell(row=3,column=1).value='Source Link'
    sheet.cell(row=4,column=1).value='Unit'
    sheet.cell(row=5,column=1).value='Geography'
    
    
    sheet.cell(row=1, column=2).value=str(columns[x])
    sheet.cell(row=2, column=2).value='AHDB - Ferilizers and Feed'
    sheet.cell(row=3,column=2).value='https://ahdb.org.uk/GB-fertiliser-prices'
    sheet.cell(row=4,column=2).value='£/tonne'
    sheet.cell(row=5,column=2).value='United Kingdom'
    wb.properties.title=columns[x]+'_AHDB - Ferilizers and Feed_United Kingdom'
    wb.save(folderName + name+'_AHDB - Ferilizers and Feed_United Kingdom.xlsx')
    wb.close()