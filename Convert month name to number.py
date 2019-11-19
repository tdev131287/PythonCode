import requests
import pandas as pd
from selenium import webdriver
import time
import os
from selenium.webdriver.chrome.options import Options
import glob
import datetime
import openpyxl

now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
folderName= datestring

folderName=os.path.join(os.path.dirname(__file__)) + '/' + datestring + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)


look_up = {'Jan': '1', 'Feb': '2', 'Mar': '3', 'Apr': '4', 'May': '5',
            'Jun': '6', 'Jul': '7', 'Aug': '8', 'Sept': '9', 'Oct': '10', 'Nov': '11', 'Dec': '12'}

url = 'https://ifcndairy.org/about-ifcn-neu/ifcn-dairy-research-center-method/'
html = requests.get(url).content

df_list = pd.read_html(html,header=0)
all_data = pd.DataFrame() 


df = df_list[-3]    
columns = list(df)
for x in range(1,len(columns)):
#    print(str(columns[x]))
    cols = [0,x]
    df1 = df[df.columns[cols]]
    df1['Month']=str(columns[x])
    df1.columns=["Year", "value", "Month"]
#    print(df1)
    all_data = all_data.append(df1)

all_data['Month_Number'] = all_data['Month'].apply(lambda x: look_up[x])
all_data["Month_Number"] = pd.to_numeric(all_data["Month_Number"])
all_data=all_data.sort_values(by=['Year','Month_Number'])
all_data=all_data[["Year", "Month","Month_Number","value"]]
all_data.to_excel(folderName+'Milk_IFCN Dairy Research Center_World.xlsx',index=False)

        ####Creating detail sheet####
wb=openpyxl.load_workbook(folderName+'Milk_IFCN Dairy Research Center_World.xlsx')
sheet=wb.create_sheet('Details')


sheet.cell(row=1, column=1).value='Commodity'
sheet.cell(row=2, column=1).value='Source'
sheet.cell(row=3,column=1).value='Source Link'
sheet.cell(row=4,column=1).value='Unit'
sheet.cell(row=5,column=1).value='Geography'
sheet.cell(row=6,column=1).value='Note'

sheet.cell(row=1, column=2).value='Milk (Solid Corrected Milk: 4.0% fat, 3.3% protein)'
sheet.cell(row=2, column=2).value='IFCN Dairy Research Center'
sheet.cell(row=3,column=2).value='https://ifcndairy.org/about-ifcn-neu/ifcn-dairy-research-center-method/'
sheet.cell(row=4,column=2).value=' USD / 100 kg'
sheet.cell(row=5,column=2).value='World'
sheet.cell(row=6,column=2).value='Combined IFCN World Milk Price Indicator illustrates the world market price level for milk. It represents the milk price a milk processor could theoretically pay to its farmers, if it was selling its products on the world spot market and producing at standardised costs. A wide range between IFCN World Milk Price Indicators indicates economic stress for specialised dairies, if their main product is trading at the lower bound of the range. It is based on the weighted average of 3 IFCN world milk price indicators: 1. SMP & butter (~32%), 2. Cheese & whey (~51%), 3. WMP (~17%), based on quarterly updated shares of the related commodities traded on the world market. In order to be able to show comparable outputs, the IFCN converts all milk with natural contents into solid corrected milk (SCM). Thus, milk outputs with 4.0% fat and 3.3% protein are generated. The formula applied is: SCM = milk production * (fat % + true protein %) / 7.3'

wb.properties.title='Milk_IFCN Dairy Research Center_World'

wb.save(folderName+'Milk_IFCN Dairy Research Center_World.xlsx')
wb.close()