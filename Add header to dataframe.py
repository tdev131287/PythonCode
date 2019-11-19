import requests
import pandas as pd
import os
from datetime import date
from dateutil.relativedelta import relativedelta
import datetime
import openpyxl
import time
#now = datetime.datetime.now()
now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
folderName= datestring

folderName=os.path.join(os.path.dirname(__file__)) + '/' + 'Downloaded files' + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)


#Petrol
#try:
html = requests.get('https://www.iocl.com/Product_PreviousPrice/PetrolPreviousPriceDynamic.aspx').content
df_list = pd.read_html(html)

df = df_list[-1]
df.columns=["Date", "Delhi", "Kolkata", "Mumbai", "Chennai"]
print(df)
#    df['Unit'] = '$ per Kg'
#    df['Source']='African Tea Brokers Ltd'
df.to_excel(folderName + 'Petrol_Indian Oil Corporation_India.xlsx',index=False)
time.sleep(3)
wb=openpyxl.load_workbook(filename=folderName +  'Petrol_Indian Oil Corporation_India.xlsx')
sheet=wb.create_sheet('Details')    
sheet.cell(row=1, column=1).value='Commodity'
sheet.cell(row=2, column=1).value='Source'
sheet.cell(row=3,column=1).value='Source Link'
sheet.cell(row=4,column=1).value='Unit'
sheet.cell(row=5,column=1).value='Geography'


sheet.cell(row=1, column=2).value='Petrol'
sheet.cell(row=2, column=2).value='Indian Oil Corporation'
sheet.cell(row=3,column=2).value='https://www.iocl.com/Product_PreviousPrice/PetrolPreviousPriceDynamic.aspx'
sheet.cell(row=4,column=2).value='INR per Liter'
sheet.cell(row=5,column=2).value='India'
wb.save(folderName + 'Petrol_Indian Oil Corporation_India.xlsx')
wb.close()
#except:
#    print('no table found')




