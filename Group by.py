import os

import glob
import requests
import json
import pandas as pd
import datetime
import time
import openpyxl
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl


now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str('{:02d}'.format(now.day))
folderName= datestring

folderName=os.path.join(os.path.dirname(__file__)) + '/' + 'Downloaded files' + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)

files=['Shell','Caltex','SPC']
#files=['Caltex']
for file in files:
    strType=""
    headers = {
        'sec-fetch-mode': 'cors',
        'origin': 'https://www.chenlim.com',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'en-IN,en-GB;q=0.9,en-US;q=0.8,en;q=0.7,hi;q=0.6',
        'x-requested-with': 'XMLHttpRequest',
        'cookie': '_ga=GA1.2.883953522.1564639334; _gid=GA1.2.339182326.1566897384; _gat=1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36',
        'content-type': 'application/x-www-form-urlencoded',
        'accept': '*/*',
        'referer': 'https://www.chenlim.com/fuel/',
        'authority': 'www.chenlim.com',
        'sec-fetch-site': 'same-origin',
    }
    
    data = {
      'action': 'getByBrand',
      'pric_brand': file,
      'dStart': ' 2019-09-20',
      'dEnd': '2019-09-20'
    }
    
    response = requests.post('https://www.chenlim.com/fuel/get_price.php', headers=headers, data=data)
    #print(response.text)
    strData=response.text
    
    vardata=strData.split('|')
    #print(vardata)
    wb=openpyxl.Workbook()
    sheet=wb['Sheet'] 
    
    sheet.cell(row=1, column=1).value='Date'    
    sheet.cell(row=1, column=2).value='Price'
    sheet.cell(row=1, column=3).value='Type'
    r=2
    print('Processing:' + file)
    for x in vardata:
    #    date=x.split('|')
#        print('#######################')
    #    print(x.split(',')[0])
#        print(x.split(','))
        try:
            
            sheet.cell(row=r, column=2).value=x.split(',')[1]
            
#            'sheet.cell(row=r, column=1).value=datetime.datetime.strptime(x.split(',')[0], '%Y-%m-%d %H:%M:%S')
            dt=datetime.datetime.strptime(x.split(',')[0], '%Y-%m-%d %H:%M:%S')
            sheet.cell(row=r, column=1).value=str(dt.year) + '-' + str('{:02d}'.format(dt.month)) +'-'+ str('{:02d}'.format(dt.day))
            sheet.cell(row=r, column=3).value=strType
#            print(datetime.datetime.strptime(x.split(',')[0], '%Y-%m-%d %H:%M:%S'))


            r=r+1
        except:
            if x!="":
                strType=str(x)
#                print(strType)
#            print(x)
#            print("##############################################")
#            pass
    wb.save(folderName + file + '_ChenLim_Singapore.xlsx')
    wb.close()
    df=pd.read_excel(folderName + file + '_ChenLim_Singapore.xlsx')
    grouped = df.groupby(['Date', 'Type']).mean()
    df1 = df.groupby(['Date', 'Type']).mean().reset_index()
    df1['Date']= pd.to_datetime(df1['Date']) 
    print(df1)
    df1.to_excel(folderName + file + '_ChenLim_Singapore.xlsx',index=False) 
    
    
    wb=openpyxl.load_workbook(folderName + file + '_ChenLim_Singapore.xlsx')
    sheet=wb.create_sheet('Details')  
    sheet.cell(row=1, column=1).value='Commodity'
    sheet.cell(row=2, column=1).value='Source'
    sheet.cell(row=3,column=1).value='Source Link'
    sheet.cell(row=4,column=1).value='Unit'
    sheet.cell(row=5,column=1).value='Geography'
    
    
    
    sheet.cell(row=1, column=2).value=file
    sheet.cell(row=2, column=2).value='ChenLim'
    sheet.cell(row=3,column=2).value='https://www.chenlim.com'
    sheet.cell(row=4,column=2).value='SGD per litre'
    sheet.cell(row=5,column=2).value='Singapore'
    
    
        
    wb.save(folderName + file + '_ChenLim_Singapore.xlsx')
    wb.close()
    

#    
    