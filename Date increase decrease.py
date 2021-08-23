import os
import pandas as pd
import datetime
import time
from selenium import webdriver
# download file
from dateutil.relativedelta import relativedelta
from datetime import date

import openpyxl

# ==================decreasing 2 days from the date==================
end_date = datetime.datetime.today() - datetime.timedelta(days=2)
end_date=end_date.date()
# ==============================================================

now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str(now.day)
folderName= datestring

now = date.today() + relativedelta(days=-2)

Par1=str(now.year) +  str('{:02d}'.format(now.month)) + str(now.day)

folderName=os.path.join(os.path.dirname(__file__)) + '/' + folderName + '/'

if not os.path.exists(folderName):
    os.mkdir(folderName)

chrome_options = webdriver.ChromeOptions()
path=folderName
path= path.replace("/","\\")
prefs = {'download.default_directory' : path}
chrome_options.add_experimental_option('prefs', prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)


   
try:  
    driver.get("http://news.agrofy.com.ar/granos/precios/series-historicas/pizarra/filtros/20100101/" + Par1 + "/4,3,5,10,1,63/10001,10002,10003,10004,10005/Pesos?descargar=1")
    time.sleep(60)
except:
    print('File not download for ')

driver.quit()


df=pd.read_excel('Mapping.xlsx',sheet_name='Sheet1')
df1=pd.read_excel(folderName + 'fyo-series-historicas.xlsx',header=3)
#df1=df1[1:]

result=df.merge(df1,on='Producto')
list1=result['Product Name'].unique().tolist()
#print(list1)

for comname in list1:
    result1 = result[result['Product Name'] == comname]
    result1=result1.sort_values('Fecha de rueda',ascending=False)
    result1.to_excel(folderName + comname + '_Agrofy News_Argentina.xlsx', index=False)    
    
    #Creating detail sheet
    wb=openpyxl.load_workbook(folderName + comname + '_Agrofy News_Argentina.xlsx')
    sheet=wb.create_sheet('Details')
    
    sheet.cell(row=1, column=1).value='Commodity'
    sheet.cell(row=2, column=1).value='Source'
    sheet.cell(row=3,column=1).value='Source Link'
    sheet.cell(row=4,column=1).value='Unit'
    sheet.cell(row=5,column=1).value='Geography'
    
    
    sheet.cell(row=1, column=2).value=comname
    sheet.cell(row=2, column=2).value='Agrofy News'
    sheet.cell(row=3,column=2).value='https://news.agrofy.com.ar/granos/precios/series-historicas/pizarra'
    sheet.cell(row=4,column=2).value='ARS per ton'
    sheet.cell(row=5,column=2).value='Argentina'
        
    wb.save(folderName + comname + '_Agrofy News_Argentina.xlsx')
    wb.close()
os.remove(folderName + 'fyo-series-historicas.xlsx')
