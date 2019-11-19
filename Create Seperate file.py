import os
import pandas as pd
import datetime
import time
from selenium import webdriver
import openpyxl
# download file

now = datetime.datetime.now()
datestring=str(now.year) + '-' + str('{:02d}'.format(now.month)) +'-'+ str(now.day)
#folderName= datestring
#
#folderName=os.path.join(os.path.dirname(__file__)) + '/' + folderName + '/'
#
#if not os.path.exists(folderName):
#    os.mkdir(folderName)

#
#
folderName=os.path.join(os.path.dirname(__file__)) + '/' + 'downloaded files' + '/'

chrome_options = webdriver.ChromeOptions()
path=folderName
path= path.replace("/","\\")
prefs = {'download.default_directory' : path}
chrome_options.add_experimental_option('prefs', prefs)
driver = webdriver.Chrome(chrome_options=chrome_options)



url='http://english.czce.com.cn/enportal/DFSStaticFiles/Future/' + str(now.year) + '/EnglishFutureDataAllHistory/ALL.xls'
try:  
    driver.get(url)
    time.sleep(20)
    os.rename(os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + str('All.xls'), os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + datestring + '.xls')
except:
    print('File not download for ' + url)


driver.quit()



df=pd.read_excel('Mapping.xlsx',sheet_name='Sheet1')
df1=pd.read_excel(os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + datestring + '.xls',sheet_name='sheet1',header=1)
#df1=df1[1:]

result=df.merge(df1,on='Contract Code')
list1=result['Commodity'].unique().tolist()
print(list1)

for comname in list1:
    result1 = result[result['Commodity'] == comname]
    result1.to_excel(os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + comname + '_Zhengzhou Commodity Exchange_China.xlsx', index=False)
    
    
    #Creating detail sheet
    wb=openpyxl.load_workbook(os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + comname + '_Zhengzhou Commodity Exchange_China.xlsx')
    sheet=wb.create_sheet('Details')
    
    sheet.cell(row=1, column=1).value='Commodity'
    sheet.cell(row=2, column=1).value='Source'
    sheet.cell(row=3,column=1).value='Source Link'
    sheet.cell(row=4,column=1).value='Unit'
    sheet.cell(row=5,column=1).value='Geography'
    
    
    sheet.cell(row=1, column=2).value=comname
    sheet.cell(row=2, column=2).value='Zhengzhou Commodity Exchange'
    sheet.cell(row=3,column=2).value='http://english.czce.com.cn/enportal/MarketData/Futures/DailyTradingData/H69030701index_1.htm'
    sheet.cell(row=4,column=2).value='Yuan/ton'
    sheet.cell(row=5,column=2).value='China'
        
    wb.save(os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + comname + '_Zhengzhou Commodity Exchange_China.xlsx')
    wb.close()
os.remove(os.path.join(os.path.dirname(__file__)) + '/downloaded files/' + datestring + '.xls')
